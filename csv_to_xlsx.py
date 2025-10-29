#!/usr/bin/env python3
"""Small utility: convert a CSV checkpoint to an XLSX file using streaming write-only workbook.

Usage:
  py csv_to_xlsx.py --input nafdac_greenbook.csv --output nafdac_greenbook.xlsx

This reads the CSV twice: first pass to compute reasonable column widths, second pass to write rows.
"""
import csv
import argparse
import os
Workbook = None



def compute_max_widths(csv_path, encoding='utf-8'):
    max_widths = []
    with open(csv_path, newline='', encoding=encoding) as f:
        reader = csv.reader(f)
        for row in reader:
            for i, cell in enumerate(row):
                l = len(str(cell)) if cell is not None else 0
                if i >= len(max_widths):
                    max_widths.append(l)
                else:
                    if l > max_widths[i]:
                        max_widths[i] = l
    return max_widths


def csv_to_xlsx(csv_path, xlsx_path, encoding='utf-8'):
    if not os.path.exists(csv_path):
        raise FileNotFoundError(f"CSV file not found: {csv_path}")

    print(f"Computing column widths for {csv_path}...")
    widths = compute_max_widths(csv_path, encoding=encoding)

    # Import openpyxl at runtime to avoid module-level import issues across different python launches
    try:
        from openpyxl import Workbook
    except Exception:
        raise RuntimeError("openpyxl is not installed. Install it with: pip install openpyxl")

    print(f"Writing to Excel: {xlsx_path} (this may take a moment)...")
    wb = Workbook(write_only=True)
    ws = wb.create_sheet(title="NAFDAC Greenbook")

    with open(csv_path, newline='', encoding=encoding) as f:
        reader = csv.reader(f)
        for row in reader:
            # Convert None to empty strings
            out = [c if c is not None else '' for c in row]
            ws.append(out)

    # Try to set column widths (openpyxl write-only supports column_dimensions assignment)
    try:
        for i, w in enumerate(widths, start=1):
            col_letter = None
            try:
                from openpyxl.utils import get_column_letter
                col_letter = get_column_letter(i)
                # add padding, cap width
                adj = min(max(10, int(w * 1.1) + 2), 60)
                ws.parent.active = ws  # ensure active
                ws.parent.column_dimensions[col_letter].width = adj
            except Exception:
                # ignore width set failures in write-only mode
                pass
    except Exception:
        pass

    wb.save(xlsx_path)
    print(f"Saved Excel file: {xlsx_path}")


def main():
    p = argparse.ArgumentParser(description="Convert CSV checkpoint to XLSX (streaming)")
    p.add_argument("--input", "-i", default="nafdac_greenbook.csv", help="Input CSV file")
    p.add_argument("--output", "-o", default="nafdac_greenbook.xlsx", help="Output XLSX file")
    args = p.parse_args()

    try:
        csv_to_xlsx(args.input, args.output)
    except Exception as e:
        print(f"Error: {e}")


if __name__ == '__main__':
    main()
