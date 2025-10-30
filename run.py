from selenium import webdriver
import argparse
import os
from selenium.webdriver.chrome.service import Service
from selenium.webdriver.common.by import By
from selenium.webdriver.support.ui import WebDriverWait
from selenium.webdriver.support import expected_conditions as EC
from selenium.common.exceptions import TimeoutException, UnexpectedAlertPresentException, NoAlertPresentException, StaleElementReferenceException, InvalidSessionIdException
import csv
import traceback
import requests
from datetime import datetime
from webdriver_manager.chrome import ChromeDriverManager
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from selenium.webdriver.common.alert import Alert
import time

# Pre-run convenience: allow setting a start page via environment variable or a small file
# Priority: CLI --start > START_PAGE env var > start_page.txt file > existing checkpoint detection
def compute_start_page_from_files(base_name="nafdac_greenbook_1400"):
    """Return an estimated start page based on existing CSV or Excel checkpoint files.
    Returns None if no checkpoint is present.
    """
    csv_path = base_name + ".csv"
    xlsx_path = base_name + ".xlsx"
    try:
        if os.path.exists(csv_path):
            with open(csv_path, newline='', encoding='utf-8') as f:
                reader = csv.reader(f)
                # skip header
                next(reader, None)
                count = sum(1 for _ in reader)
            return (count // 10) + 1 if count else 1
        if os.path.exists(xlsx_path):
            try:
                from openpyxl import load_workbook
                wb = load_workbook(xlsx_path, read_only=True)
                ws = wb.active
                # subtract header
                count = 0
                for _ in ws.iter_rows(min_row=2, values_only=True):
                    count += 1
                return (count // 10) + 1 if count else 1
            except Exception:
                return None
    except Exception:
        return None
    return None

def save_to_excel(data, output_file):
    import os
    from datetime import datetime
    
    # Try to save to the main file first
    def try_save(file_path, workbook):
        try:
            workbook.save(file_path)
            return True
        except PermissionError:
            return False
            
    # Create a new workbook and select active sheet
    wb = Workbook()
    ws = wb.active
    ws.title = "NAFDAC Greenbook"
    
    # Define headers
    headers = [
        "Product Name", "Active Ingredient", "Dosage Form",
        "Product Category", "NAFDAC Reg No", "Applicant", 
        "Manufacturer", "Approval Date"
    ]
    
    # Write headers with styling
    header_font = Font(bold=True, color="FFFFFF")
    header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
    header_alignment = Alignment(horizontal="center", vertical="center")
    border = Border(
        left=Side(style='thin'), 
        right=Side(style='thin'), 
        top=Side(style='thin'), 
        bottom=Side(style='thin')
    )
    
    for col, header in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col, value=header)
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = header_alignment
        cell.border = border
        
    # Write data
    for row_idx, row_data in enumerate(data, 2):
        for col_idx, value in enumerate(row_data, 1):
            cell = ws.cell(row=row_idx, column=col_idx, value=value)
            cell.border = border
            cell.alignment = Alignment(vertical="center")
    
    # Adjust column widths
    for col in range(1, len(headers) + 1):
        column_letter = get_column_letter(col)
        max_length = 0
        column = ws[column_letter]
        for cell in column:
            try:
                if len(str(cell.value)) > max_length:
                    max_length = len(str(cell.value))
            except:
                pass
        adjusted_width = (max_length + 2) * 1.2
        ws.column_dimensions[column_letter].width = min(adjusted_width, 50)
    
    # Try to save the workbook with retry mechanism
    if not try_save(output_file, wb):
        # If main file is locked, create a backup file
        backup_file = f"nafdac_greenbook_backup_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"
        print(f"Main file is locked. Saving to backup file: {backup_file}")
        if try_save(backup_file, wb):
            print("Successfully saved to backup file")
        else:
            # If both fails, try saving to temp directory
            temp_file = os.path.join(os.environ.get('TEMP', ''), backup_file)
            if try_save(temp_file, wb):
                print(f"Saved to temporary location: {temp_file}")
            else:
                raise Exception("Could not save file in any location")


def append_rows_to_csv(rows, csv_path):
    """Append rows (list of lists) to a CSV file. Creates file with header if missing."""
    if not rows:
        return
    header = [
        "Product Name", "Active Ingredient", "Dosage Form",
        "Product Category", "NAFDAC Reg No", "Applicant",
        "Manufacturer", "Approval Date"
    ]
    exists = os.path.exists(csv_path)
    with open(csv_path, "a", newline='', encoding='utf-8') as f:
        writer = csv.writer(f)
        if not exists:
            writer.writerow(header)
        for r in rows:
            writer.writerow(r)

def load_existing_data(output_file):
    # Prefer CSV checkpoint if available for faster/resilient resume
    csv_file = os.path.splitext(output_file)[0] + ".csv"
    data = []
    try:
        if os.path.exists(csv_file):
            with open(csv_file, newline='', encoding='utf-8') as f:
                reader = csv.reader(f)
                # skip header
                headers = next(reader, None)
                for row in reader:
                    if any(cell.strip() for cell in row if isinstance(cell, str)):
                        data.append(row)
            last_page = (len(data) // 10) + 1 if data else 1
            print(f"Loaded {len(data)} existing records from CSV (approximately page {last_page})")
            return data, last_page

        # Fallback to Excel load if CSV missing
        from openpyxl import load_workbook
        wb = load_workbook(output_file)
        ws = wb.active
        for row in ws.iter_rows(min_row=2, values_only=True):  # Skip header row
            if any(cell is not None for cell in row):  # Only include non-empty rows
                data.append(list(row))

        # Calculate the page number based on records (10 records per page)
        last_page = (len(data) // 10) + 1 if data else 1
        print(f"Loaded {len(data)} existing records (approximately page {last_page})")
        return data, last_page
    except Exception as e:
        print(f"No existing data found or error loading file: {e}")
        return [], 1


def log_skipped_page(page, reason, log_path="skipped_pages.log"):
    try:
        ts = datetime.now().strftime('%Y-%m-%d %H:%M:%S')
        with open(log_path, 'a', encoding='utf-8') as f:
            f.write(f"{ts}\tpage:{page}\treason:{reason}\n")
    except Exception as e:
        print(f"Failed to write skip log: {e}")


def csv_to_excel_stream(csv_path, excel_path):
    """Convert CSV to XLSX using openpyxl write-only workbook to avoid heavy memory/time cost."""
    from openpyxl import Workbook
    wb = Workbook(write_only=True)
    ws = wb.create_sheet(title="NAFDAC Greenbook")

    # Read CSV and write rows directly
    with open(csv_path, newline='', encoding='utf-8') as f:
        reader = csv.reader(f)
        for r in reader:
            ws.append(r)

    # Ensure parent dir exists
    try:
        wb.save(excel_path)
        print(f"Converted CSV checkpoint to Excel: {excel_path}")
    except Exception as e:
        print(f"Failed to write Excel from CSV: {e}")


def detect_datatables_ajax(driver):
    """Return the DataTables ajax URL (string) and info dict if the table is serverSide, else (None, None)."""
    try:
        js = ("return (function(){"
              "var tbl = document.querySelector('table.dataTable');"
              "if(!tbl) return null;"
              "var dt=null;"
              "try{ dt = window.jQuery && jQuery(tbl).DataTable ? jQuery(tbl).DataTable() : null; }catch(e){ dt=null; }"
              "if(!dt){ try{ dt = $.fn.dataTable.Api(tbl); }catch(e){ dt=null; } }"
              "if(!dt) return null;"
              "var settings = null; try{ settings = dt.page.info ? dt.page.info() : (dt.settings ? dt.settings()[0] : null);}catch(e){ settings=null;}"
              "var ajax = null; try{ ajax = dt.ajax ? dt.ajax : (dt.settings && dt.settings()[0] && dt.settings()[0].oFeatures ? dt.settings()[0].ajax : null);}catch(e){ ajax=null;}"
              "if(ajax && typeof ajax === 'object' && ajax.url) return {ajax: ajax.url, info: settings};"
              "if(ajax && typeof ajax === 'string') return {ajax: ajax, info: settings};"
              "return {ajax:null, info:settings}; })();")
        res = driver.execute_script(js)
        if not res:
            return None, None
        ajax = res.get('ajax') if isinstance(res, dict) else None
        info = res.get('info') if isinstance(res, dict) else None
        return ajax, info
    except Exception as e:
        print(f"detect_datatables_ajax error: {e}")
        return None, None


def detect_ajax_from_html(html, base_url="https://greenbook.nafdac.gov.ng/"):
    """Try to find a DataTables ajax URL in the HTML (fallback for --force-api).
    Returns absolute URL or None."""
    import re
    # common patterns: "ajax": "URL"  or ajax: { url: 'URL' }
    patterns = [r'"ajax"\s*:\s*"([^"]+)"', r"ajax\s*:\s*\{\s*url\s*:\s*'([^']+)'",
                r"ajax\s*:\s*'([^']+)'", r'ajax\s*:\s*\{\s*url\s*:\s*\"([^\"]+)\"']
    for p in patterns:
        m = re.search(p, html)
        if m:
            url = m.group(1)
            if url.startswith('http'):
                return url
            # relative -> join with base
            from urllib.parse import urljoin
            return urljoin(base_url, url)
    return None


def api_scrape(ajax_url, start_page, end_page, csv_checkpoint, cookies=None, headers=None, page_length=10):
    """Scrape pages via the DataTables server-side AJAX endpoint and append rows to CSV checkpoint.
    Returns total rows scraped.
    """
    s = requests.Session()
    if headers:
        s.headers.update(headers)
    if cookies:
        for c in cookies:
            s.cookies.set(c['name'], c.get('value',''))

    total_rows = 0
    for page in range(start_page, end_page + 1):
        start = (page - 1) * page_length
        params = {
            'start': start,
            'length': page_length,
            'draw': page
        }
        try:
            resp = s.get(ajax_url, params=params, timeout=30)
            resp.raise_for_status()
            j = resp.json()
            data = None
            if isinstance(j, dict):
                data = j.get('data') or j.get('aaData')
            elif isinstance(j, list):
                data = j
            if not data:
                print(f"No data returned for page {page} (status {resp.status_code})")
                break

            # Normalize rows: if rows are dicts, take values; if lists, use directly
            page_rows = []
            for row in data:
                if isinstance(row, dict):
                    # preserve order by keys if possible, else values()
                    vals = list(row.values())
                    page_rows.append(vals)
                elif isinstance(row, (list, tuple)):
                    page_rows.append(list(row))
                else:
                    page_rows.append([str(row)])

            append_rows_to_csv(page_rows, csv_checkpoint)
            total_rows += len(page_rows)
            print(f"API scraped page {page}: {len(page_rows)} rows (total {total_rows})")

            # If fewer than page_length returned, probably last page
            if len(page_rows) < page_length:
                break

        except Exception as e:
            print(f"API scraping error on page {page}: {e}")
            break

    return total_rows


def api_get_page(ajax_url, page, cookies=None, headers=None, page_length=10):
    """Return the raw data array for a given 1-based page from the DataTables ajax endpoint."""
    s = requests.Session()
    if headers:
        s.headers.update(headers)
    if cookies:
        for c in cookies:
            s.cookies.set(c['name'], c.get('value',''))

    start = (page - 1) * page_length
    params = {'start': start, 'length': page_length, 'draw': page}
    resp = s.get(ajax_url, params=params, timeout=30)
    resp.raise_for_status()
    j = resp.json()
    if isinstance(j, dict):
        return j.get('data') or j.get('aaData') or []
    if isinstance(j, list):
        return j
    return []


def find_resume_page_via_api(last_identifier, ajax_url, est_page, cookies=None, headers=None, page_length=10, id_index=4, max_scan=50):
    """Scan nearby pages (starting at est_page) to find the page that contains last_identifier.
    Returns the page number that contains it, or est_page if not found.
    id_index: column index in returned row data that contains the unique identifier (NAFDAC Reg No)
    max_scan: how many pages to scan forward/backward before giving up
    """
    try:
        # First check estimated page
        pages_to_check = [est_page]
        # then expand forwards and backwards alternately
        for i in range(1, max_scan+1):
            pages_to_check.append(est_page + i)
            if est_page - i > 0:
                pages_to_check.append(est_page - i)

        checked = set()
        for p in pages_to_check:
            if p in checked or p < 1:
                continue
            checked.add(p)
            try:
                rows = api_get_page(ajax_url, p, cookies=cookies, headers=headers, page_length=page_length)
            except Exception:
                continue
            for r in rows:
                # try to get identifier
                try:
                    if isinstance(r, dict):
                        vals = list(r.values())
                    else:
                        vals = list(r)
                    ident = vals[id_index] if len(vals) > id_index else None
                    if ident and str(ident).strip() == str(last_identifier).strip():
                        return p
                except Exception:
                    continue
        return est_page
    except Exception:
        return est_page

def scrape_greenbook(output_file="nafdac_greenbook.xlsx", end_page=876, resume=True, driver_path=None, start_page=None, no_headless=False, debug=False, force_api=False, csv_only=False):
    # Setup Chrome options
    options = webdriver.ChromeOptions()
    if not no_headless:
        options.add_argument("--headless=new")
    options.add_argument("--no-sandbox")
    options.add_argument("--disable-dev-shm-usage")
    options.add_argument("--window-size=1920,1080")  # Larger window for better rendering
    
    def init_driver():
        """Initialize and return a Chrome webdriver instance (retries on failures)."""
        last_err = None
        for attempt in range(3):
            try:
                if driver_path:
                    d = webdriver.Chrome(service=Service(driver_path), options=options)
                else:
                    d = webdriver.Chrome(service=Service(ChromeDriverManager().install()), options=options)
                return d
            except Exception as e:
                last_err = e
                print(f"Driver init failed (attempt {attempt+1}/3): {e}")
                time.sleep(2)
        raise Exception(f"Could not initialize Chrome driver: {last_err}")

    # If user requested force-api, try to detect the AJAX endpoint from the HTML and run the API scraper
    # without initializing Selenium (this avoids webdriver_manager probing the local browser).
    if force_api:
        try:
            csv_checkpoint = os.path.splitext(output_file)[0] + ".csv"
            # Determine start page from existing checkpoint or provided start_page
            data, last_page = load_existing_data(output_file)
            if start_page:
                page = int(start_page)
            else:
                page = last_page if resume else 1
            print(f"(force-api) starting from page {page}")

            main_html = None
            try:
                main_html = requests.get('https://greenbook.nafdac.gov.ng/', timeout=20).text
            except Exception as rexc:
                print(f"Failed to fetch main page for force-api detection: {rexc}")

            force_ajax = None
            if main_html:
                force_ajax = detect_ajax_from_html(main_html, base_url='https://greenbook.nafdac.gov.ng/')

            if force_ajax:
                print(f"Force-API detected ajax endpoint from HTML: {force_ajax}")
                headers = {'User-Agent': 'Mozilla/5.0', 'Referer': 'https://greenbook.nafdac.gov.ng/'}
                scraped = api_scrape(force_ajax, page, end_page, csv_checkpoint, headers=headers)
                print(f"Force-API scraping finished, {scraped} rows appended to {csv_checkpoint}")
                return
            else:
                print("--force-api requested but could not detect ajax endpoint from HTML; falling back to Selenium")
        except Exception as e:
            print(f"Force-API early path failed: {e}; continuing with Selenium")

    # Initialize driver
    driver = init_driver()
    
    try:
        # Load the page
        print("Loading page...")
        driver.get("https://greenbook.nafdac.gov.ng/")
        
        # Wait for table to load
        wait = WebDriverWait(driver, 20)
        print("Waiting for table to appear...")
        wait.until(EC.presence_of_element_located((By.CSS_SELECTOR, "table.dataTable")))
        time.sleep(2)  # Allow table to fully render

        # If debug mode is on, dump DataTables info to help debugging JS API issues
        if debug:
            try:
                print("--- DEBUG: DataTables inspection ---")
                js = ("return (function(){"
                      "var res=[];"
                      "var tables = document.querySelectorAll('table.dataTable');"
                      "for (var i=0;i<tables.length;i++){"
                      " var tbl = tables[i];"
                      " var info={};"
                      " info.index=i; info.id=tbl.id||null;"
                      " try{ info.html = tbl.outerHTML.substring(0,500); }catch(e){ info.html='err'; }"
                      " try{ if(window.jQuery && jQuery(tbl).DataTable) var dt = jQuery(tbl).DataTable(); else if(window.jQuery && jQuery.fn && jQuery.fn.dataTable && jQuery.fn.dataTable.Api) var dt = new jQuery.fn.dataTable.Api(tbl); else var dt = null;}catch(e){ dt=null; info.dt_err = e.toString(); }"
                      " if(dt){ try{ info.page = dt.page(); }catch(e){ info.page='err'; } try{ info.info = dt.page.info(); }catch(e){ info.info='err'; } } else { info.dt='no-dt'; }"
                      " res.push(info); } return res; })();")
                dt_info = driver.execute_script(js)
                print(dt_info)
                print("--- END DEBUG ---")
            except Exception as e:
                print(f"Debug dump failed: {e}")

        # Detect if DataTables is server-side and has an AJAX endpoint. If so, prefer API scraping.
        try:
            ajax_url, dt_info = detect_datatables_ajax(driver)
            if ajax_url and dt_info and dt_info.get('serverSide'):
                print(f"Detected DataTables server-side ajax: {ajax_url} (info: {dt_info})")
                # Prepare headers/cookies for requests
                headers = {'User-Agent': 'Mozilla/5.0', 'Referer': driver.current_url}
                cookies = driver.get_cookies()
                csv_checkpoint = os.path.splitext(output_file)[0] + ".csv"
                scraped = api_scrape(ajax_url, page, end_page, csv_checkpoint, cookies=cookies, headers=headers, page_length=dt_info.get('length',10))
                print(f"API-mode scraping finished, {scraped} rows appended to {csv_checkpoint}")
                # Close the browser and exit early since we've completed via API
                driver.quit()
                return
        except Exception as e:
            print(f"API detection/scrape skipped due to error: {e}")

        # If user requested force-api, attempt to detect AJAX endpoint from HTML and run api_scrape without Selenium
        if force_api:
            try:
                main_html = None
                try:
                    main_html = requests.get('https://greenbook.nafdac.gov.ng/', timeout=20).text
                except Exception as rexc:
                    print(f"Failed to fetch main page for API-detection: {rexc}")

                force_ajax = None
                if main_html:
                    force_ajax = detect_ajax_from_html(main_html, base_url='https://greenbook.nafdac.gov.ng/')

                if force_ajax:
                    print(f"Force-API detected ajax endpoint: {force_ajax}")
                    csv_checkpoint = os.path.splitext(output_file)[0] + ".csv"
                    scraped = api_scrape(force_ajax, page, end_page, csv_checkpoint, headers={'User-Agent':'Mozilla/5.0','Referer':'https://greenbook.nafdac.gov.ng/'})
                    print(f"Force-API scraping finished, {scraped} rows appended to {csv_checkpoint}")
                    try:
                        driver.quit()
                    except:
                        pass
                    return
                else:
                    print("--force-api requested but could not detect ajax endpoint from HTML; continuing with Selenium")
            except Exception as e:
                print(f"Force-API path failed: {e}")
        
        # Load existing data and determine start page
        data, last_page = load_existing_data(output_file)
        if start_page:
            page = int(start_page)
            print(f"Starting from explicit start page {page}")
        else:
            page = last_page if resume else 1
        print(f"Starting from page {page}")

        # If we have existing data, try to locate the exact page that contains the last row
        # so we can resume from the next page instead of re-scraping pages or starting at 1.
        if data:
            last_row = data[-1]
            last_identifier = None
            try:
                # Prefer NAFDAC Reg No column if present (index 4)
                if len(last_row) > 4:
                    last_identifier = last_row[4]
                else:
                    # fallback to last column
                    last_identifier = last_row[-1]
            except Exception:
                last_identifier = None

            if last_identifier:
                print(f"Last scraped identifier from checkpoint: {last_identifier}")
                # Try to detect ajax endpoint for fast lookup
                try:
                    ajax_url2, dt_info2 = detect_datatables_ajax(driver)
                except Exception:
                    ajax_url2, dt_info2 = (None, None)

                if ajax_url2:
                    try:
                        est = page
                        found_page = find_resume_page_via_api(last_identifier, ajax_url2, est_page=est, cookies=driver.get_cookies(), headers={'User-Agent':'Mozilla/5.0','Referer':driver.current_url}, page_length=(dt_info2.get('length',10) if dt_info2 else 10))
                        if found_page:
                            # resume from the page after the one containing last_identifier
                            page = found_page + 1
                            print(f"Detected last identifier on page {found_page} via API — resuming from {page}")
                    except Exception as e:
                        print(f"API resume detection failed: {e}")
                elif force_api:
                    # Attempt HTML detection if user explicitly requested force-api
                    try:
                        main_html = requests.get('https://greenbook.nafdac.gov.ng/', timeout=20).text
                        force_ajax2 = detect_ajax_from_html(main_html, base_url='https://greenbook.nafdac.gov.ng/')
                        if force_ajax2:
                            est = page
                            found_page = find_resume_page_via_api(last_identifier, force_ajax2, est_page=est, headers={'User-Agent':'Mozilla/5.0','Referer':'https://greenbook.nafdac.gov.ng/'}, page_length=10)
                            page = found_page + 1
                            print(f"Detected last identifier on page {found_page} via forced-API HTML detection — resuming from {page}")
                    except Exception as e:
                        print(f"Force-API resume detection failed: {e}")
                else:
                    # Fall back to a Selenium-based localized search: jump to estimated page and verify
                    try:
                        est = page
                        print(f"Attempting to verify last scraped identifier via Selenium starting at est page {est}")
                        # Try to jump to estimated page
                        try:
                            if est > 1:
                                # Attempt a direct DataTables JS jump (0-based index)
                                js_jump = ("(function(){var tblEl = document.querySelector('table.dataTable'); if(!tblEl) return 'no-table'; var dt=null; try{ dt = $(tblEl).DataTable(); }catch(e){} if(!dt){ try{ dt = $.fn.dataTable.Api(tblEl); }catch(e){} } if(!dt) return 'no-dt'; try{ dt.page(%d).draw(false); return 'ok'; }catch(e){ return 'err'; }})();") % (est - 1)
                                try:
                                    driver.execute_script(js_jump)
                                except Exception:
                                    pass
                                time.sleep(1)
                        except Exception:
                            pass

                        def page_contains_identifier(idval):
                            try:
                                rows = driver.find_elements(By.CSS_SELECTOR, "table.dataTable tbody tr")
                                for r in rows:
                                    try:
                                        cells = r.find_elements(By.TAG_NAME, 'td')
                                        vals = [c.text.strip() for c in cells]
                                        if len(vals) > 4 and vals[4].strip() == str(idval).strip():
                                            return True
                                    except Exception:
                                        continue
                            except Exception:
                                return False
                            return False

                        found = False
                        if page_contains_identifier(last_identifier):
                            found = True
                            found_page = est
                        else:
                            # scan a few pages forward (server-side ordering is stable)
                            for offset in range(1, 6):
                                try:
                                    target = est + offset
                                    # Try DataTables JS jump to target
                                    js_jump2 = ("(function(){var tblEl = document.querySelector('table.dataTable'); if(!tblEl) return 'no-table'; var dt=null; try{ dt = $(tblEl).DataTable(); }catch(e){} if(!dt){ try{ dt = $.fn.dataTable.Api(tblEl); }catch(e){} } if(!dt) return 'no-dt'; try{ dt.page(%d).draw(false); return 'ok'; }catch(e){ return 'err'; }})();") % (target - 1)
                                    try:
                                        driver.execute_script(js_jump2)
                                    except Exception:
                                        pass
                                    time.sleep(1)
                                    if page_contains_identifier(last_identifier):
                                        found = True
                                        found_page = target
                                        break
                                except Exception:
                                    break

                        if found:
                            page = found_page + 1
                            print(f"Found last identifier on page {found_page} via Selenium — resuming from {page}")
                        else:
                            print("Could not locate exact last identifier via Selenium — falling back to estimated page")
                    except Exception as e:
                        print(f"Selenium resume detection failed: {e}")
        
        # Navigate to start page
        if page > 1:
            print(f"Navigating to page {page}...")
            current_page = 1
            
            def handle_alerts():
                try:
                    alert = Alert(driver)
                    alert.accept()
                    print("Cleared alert during navigation")
                    time.sleep(1)
                except NoAlertPresentException:
                    pass

            def safe_click_next():
                handle_alerts()
                next_button = wait.until(EC.element_to_be_clickable((By.CSS_SELECTOR, "li.page-item.next:not(.disabled) a")))
                driver.execute_script("arguments[0].scrollIntoView({block: 'center'});", next_button)
                time.sleep(1)
                try:
                    driver.execute_script("arguments[0].click();", next_button)
                except:
                    handle_alerts()
                    driver.execute_script("arguments[0].click();", next_button)
                time.sleep(2)  # Increased wait time after click
                handle_alerts()

            def go_to_page(target_page):
                """Try to jump to `target_page` using the DataTables JS API. Returns True on success."""
                for attempt in range(3):
                    try:
                        # Clear any alerts first
                        try:
                            Alert(driver).accept()
                        except:
                            pass

                        # Try to call DataTables API via JS. Use page index (0-based).
                        js = ("(function(){"
                              "var tblEl = document.querySelector('table.dataTable');"
                              "if(!tblEl) return 'no-table';"
                              "var dt = null;"
                              "try{ dt = $(tblEl).DataTable(); }catch(e){}"
                              "if(!dt){ try{ dt = $.fn.dataTable.Api(tblEl); }catch(e){} }"
                              "if(!dt) return 'no-dt';"
                              "try{ dt.page(%d).draw(false); return 'ok'; }catch(e){ return 'err'; }"
                              "})();") % (target_page - 1)

                        # Execute JS and then verify whether the active page becomes target_page.
                        try:
                            res = driver.execute_script(js)
                        except InvalidSessionIdException:
                            raise
                        except Exception as e:
                            res = None
                            print(f"go_to_page JS exec error: {e}")

                        # Give the table time to redraw
                        time.sleep(2)

                        # Check active page number explicitly
                        try:
                            active_el = driver.find_element(By.CSS_SELECTOR, "li.page-item.active a.page-link")
                            if active_el and active_el.text.strip().isdigit() and int(active_el.text.strip()) == target_page:
                                return True
                        except Exception:
                            pass

                        # If JS returned a diagnostic string, log it
                        if isinstance(res, str):
                            print(f"DataTables jump returned: {res}")

                        # One last short wait and re-check
                        time.sleep(1)
                        try:
                            active_el = driver.find_element(By.CSS_SELECTOR, "li.page-item.active a.page-link")
                            if active_el and active_el.text.strip().isdigit() and int(active_el.text.strip()) == target_page:
                                return True
                        except Exception:
                            pass

                        # Not successful this attempt
                        print(f"DataTables jump did not move to page {target_page} (JS res={res})")

                    except UnexpectedAlertPresentException:
                        try:
                            Alert(driver).accept()
                        except:
                            pass
                        time.sleep(1)
                    except InvalidSessionIdException:
                        # Driver lost session — propagate so caller can re-init
                        raise
                    except Exception as e:
                        print(f"go_to_page attempt error: {e}")
                        time.sleep(1)
                return False
                
            # Track repeated failures to make a stronger recovery if pagination gets stuck
            stuck_attempts = 0
            while current_page < page:
                try:
                    # Wait for navigation elements to be present
                    wait.until(EC.presence_of_element_located((By.CSS_SELECTOR, "ul.pagination")))
                    time.sleep(1)  # Give extra time for pagination to stabilize
                    
                    # Get current visible page number
                    active_page = driver.find_element(By.CSS_SELECTOR, "li.page-item.active a.page-link")
                    current_page = int(active_page.text)
                    print(f"Currently at page {current_page}")

                    # Try DataTables API jump directly to the target page (fast) as first attempt
                    try:
                        if go_to_page(page):
                            print(f"Jumped directly to page {page} using DataTables API")
                            current_page = page
                            break
                    except Exception:
                        pass
                    
                    # Try to find the last visible page number button
                    page_buttons = driver.find_elements(By.CSS_SELECTOR, "li.page-item a.page-link")
                    page_numbers = []
                    
                    for btn in page_buttons:
                        try:
                            if btn.text.strip().isdigit():
                                page_numbers.append(int(btn.text.strip()))
                        except:
                            continue
                    
                    if page_numbers:
                        # Find the highest available page number we can click
                        available_targets = [n for n in page_numbers if n > current_page]
                        if available_targets:
                            target = min(available_targets)  # Take the next available page
                            print(f"Found clickable page {target}")
                            
                            # Find and click the target page button
                            for btn in page_buttons:
                                if btn.text.strip().isdigit() and int(btn.text.strip()) == target:
                                    print(f"Clicking page {target} button...")
                                    driver.execute_script("arguments[0].scrollIntoView({block: 'center'});", btn)
                                    time.sleep(0.5)

                                    # Try multiple click/dispatch methods. After each attempt, verify the active page.
                                    clicked = False
                                    for click_attempt in range(4):
                                        try:
                                            if click_attempt == 0:
                                                driver.execute_script("arguments[0].click();", btn)
                                            elif click_attempt == 1:
                                                driver.execute_script(
                                                    "var ev = new MouseEvent('click', {bubbles: true, cancelable: true, view: window}); arguments[0].dispatchEvent(ev);",
                                                    btn)
                                            elif click_attempt == 2:
                                                # Try jQuery trigger if available
                                                try:
                                                    driver.execute_script("if(window.jQuery){ jQuery(arguments[0]).trigger('click'); }", btn)
                                                except Exception:
                                                    pass
                                            else:
                                                # Fallback: use DataTables API jump
                                                try:
                                                    if go_to_page(target):
                                                        clicked = True
                                                        break
                                                except InvalidSessionIdException:
                                                    raise

                                            time.sleep(1.5)
                                            handle_alerts()

                                            # Verify the page changed
                                            try:
                                                cur = driver.find_element(By.CSS_SELECTOR, "li.page-item.active a.page-link").text
                                                if cur.strip().isdigit() and int(cur) == target:
                                                    clicked = True
                                                    break
                                            except Exception:
                                                # As a looser fallback, check for presence of rows (table refreshed)
                                                try:
                                                    wait.until(EC.presence_of_element_located((By.CSS_SELECTOR, "table.dataTable tbody tr")))
                                                except Exception:
                                                    pass

                                        except InvalidSessionIdException:
                                            raise
                                        except Exception as click_err:
                                            print(f"Click attempt {click_attempt+1} failed: {click_err}")
                                            handle_alerts()
                                            time.sleep(1)

                                    if not clicked:
                                        print(f"Failed to navigate to page {target} via button clicks; trying DataTables API fallback")
                                        stuck_attempts += 1
                                        # If we've been stuck clicking the same pages repeatedly, try a stronger recovery
                                        if stuck_attempts >= 5:
                                            print("Pagination appears stuck — refreshing and attempting to recover")
                                            try:
                                                driver.refresh()
                                                time.sleep(3)
                                                wait.until(EC.presence_of_element_located((By.CSS_SELECTOR, "table.dataTable")))
                                                # Try to jump directly to desired page after refresh
                                                if go_to_page(page):
                                                    print(f"Recovered and jumped to page {page} after refresh")
                                                    current_page = page
                                                    break
                                            except Exception as re:
                                                print(f"Refresh recovery failed: {re}")
                                                # Try full driver restart
                                                try:
                                                    try:
                                                        driver.quit()
                                                    except:
                                                        pass
                                                    driver = init_driver()
                                                    driver.get("https://greenbook.nafdac.gov.ng/")
                                                    wait = WebDriverWait(driver, 20)
                                                    wait.until(EC.presence_of_element_located((By.CSS_SELECTOR, "table.dataTable")))
                                                    time.sleep(2)
                                                    if go_to_page(page):
                                                        print(f"Recovered and jumped to page {page} after driver restart")
                                                        current_page = page
                                                        stuck_attempts = 0
                                                        break
                                                except Exception as re2:
                                                    print(f"Driver restart recovery failed: {re2}")
                                                    return
                                        try:
                                            if go_to_page(target):
                                                print(f"Jumped to page {target} using DataTables API fallback")
                                                current_page = target
                                                break
                                        except InvalidSessionIdException:
                                            raise

                                    if clicked:
                                        current_page = target
                                    break
                            continue  # Skip the next button click for this iteration
                    
                    # If no page numbers were clickable or found, use next button
                    print(f"Using next button to advance from page {current_page}...")
                    safe_click_next()
                    
                    # Verify the page changed
                    wait.until(EC.staleness_of(active_page))
                    wait.until(EC.presence_of_element_located((By.CSS_SELECTOR, "table.dataTable tbody tr")))
                    
                    # Small pause to let the page stabilize
                    time.sleep(1)
                        
                except Exception as e:
                    print(f"Error during navigation: {e}")
                    handle_alerts()
                    time.sleep(2)
                    
                    try:
                        # Get current page after error
                        active_page = driver.find_element(By.CSS_SELECTOR, "li.page-item.active a.page-link")
                        current_page = int(active_page.text)
                        print(f"After error recovery, we are on page {current_page}")
                        
                        if current_page < page:
                            # Refresh the page if we encounter persistent errors
                            print("Refreshing page...")
                            driver.refresh()
                            time.sleep(3)
                            wait.until(EC.presence_of_element_located((By.CSS_SELECTOR, "table.dataTable")))
                        else:
                            print("Reached target page despite error")
                            break
                            
                    except:
                        # If session invalid, try to restart driver and return to the last known page
                        err = traceback.format_exc()
                        print(f"Failed to recover: {err}")
                        print("Attempting to restart the browser and resume from current target page...")
                        try:
                            try:
                                driver.quit()
                            except:
                                pass
                            driver = init_driver()
                            driver.get("https://greenbook.nafdac.gov.ng/")
                            wait = WebDriverWait(driver, 20)
                            wait.until(EC.presence_of_element_located((By.CSS_SELECTOR, "table.dataTable")))
                            time.sleep(2)
                            # Try jumping to the requested page
                            if go_to_page(page):
                                print(f"Resumed at page {page} after driver restart")
                                continue
                        except Exception as re:
                            print(f"Could not restart and resume: {re}")
                            return
        
        # Per-page failure tracking so we can skip problematic pages
        page_failures = {}

        while page <= end_page:
            print(f"Scraping page {page}...")
            try:
                # Get rows from current page
                rows = driver.find_elements(By.CSS_SELECTOR, "table.dataTable tbody tr")
                print(f"Found {len(rows)} rows on page {page}")

                # Extract data from rows
                page_data = []
                for row in rows:
                    try:
                        cells = row.find_elements(By.TAG_NAME, "td")
                        row_data = [cell.text.strip() for cell in cells]
                        if row_data:
                            page_data.append(row_data)
                    except Exception:
                        continue

                # Add to main data list
                # Add to main data list
                data.extend(page_data)
                print(f"Total records collected: {len(data)}")

                # Append this page's rows to CSV checkpoint (fast and robust)
                csv_checkpoint = os.path.splitext(output_file)[0] + ".csv"
                try:
                    append_rows_to_csv(page_data, csv_checkpoint)
                except Exception as e:
                    print(f"Warning: failed to append to CSV checkpoint: {e}")

                # Save Excel checkpoint less frequently (lighter schedule)
                if page % 50 == 0:
                    print(f"Saving Excel checkpoint at page {page}...")
                    try:
                        save_to_excel(data, output_file)
                    except Exception as e:
                        print(f"Warning: failed to save Excel checkpoint: {e}")

            except Exception as e:
                # Record failure for this page and try to recover/skip
                fail_count = page_failures.get(page, 0) + 1
                page_failures[page] = fail_count
                print(f"Error scraping page {page}: {e} (failure {fail_count})")

                # If driver session is invalid, try to restart and resume
                if isinstance(e, InvalidSessionIdException) or 'invalid session' in str(e).lower():
                    print("Detected invalid session during page scrape — restarting browser")
                    try:
                        try:
                            driver.quit()
                        except:
                            pass
                        driver = init_driver()
                        driver.get("https://greenbook.nafdac.gov.ng/")
                        wait = WebDriverWait(driver, 20)
                        wait.until(EC.presence_of_element_located((By.CSS_SELECTOR, "table.dataTable")))
                        time.sleep(2)
                        if go_to_page(page):
                            print(f"Resumed at page {page} after driver restart")
                            continue
                    except Exception as re:
                        print(f"Failed to restart driver: {re}")
                        return

                # Accept any alert and try to continue
                try:
                    Alert(driver).accept()
                    print("Accepted alert during page scrape")
                except Exception:
                    pass

                # If failures exceed threshold, skip this page
                if fail_count >= 3:
                    print(f"Page {page} failing repeatedly — skipping to next page")
                    try:
                        log_skipped_page(page, str(e))
                    except Exception:
                        pass
                    # Try to jump to the next page using DataTables API; fallback to clicking next
                    try:
                        if go_to_page(page + 1):
                            page += 1
                            continue
                    except Exception:
                        pass

                    # Try clicking next if available
                    try:
                        nb = driver.find_element(By.CSS_SELECTOR, "li.page-item.next:not(.disabled) a")
                        try:
                            driver.execute_script("arguments[0].click();", nb)
                            time.sleep(1)
                            page += 1
                            continue
                        except Exception:
                            pass
                    except Exception:
                        pass

                    # If all else fails, increment page number (best-effort skip)
                    page += 1
                    continue
            
            # Look for next button
            try:
                next_button = driver.find_element(By.CSS_SELECTOR, "li.page-item.next:not(.disabled) a")
            except:
                # Try using DataTables API to go to next page as a fallback
                print("Next button not found — attempting DataTables API jump as fallback...")
                try:
                    if go_to_page(page + 1):
                        page += 1
                        continue
                    else:
                        print("DataTables API jump failed. No more pages found.")
                        break
                except Exception as e:
                    print(f"Fallback DataTables jump error: {e}\nNo more pages found.")
                    break
                
            # Click next button and wait for table update
            max_retries = 3
            retry_count = 0
            while retry_count < max_retries:
                try:
                    # Handle any existing alerts before proceeding
                    try:
                        alert = Alert(driver)
                        alert.accept()
                        print("Cleared existing alert")
                    except NoAlertPresentException:
                        pass

                    driver.execute_script("arguments[0].scrollIntoView({block: 'center'});", next_button)
                    time.sleep(1)  # Increased pause before click
                    
                    # Click using JavaScript to avoid potential intercepted click
                    # click and wait until the active page number changes (more reliable than staleness_of)
                    try:
                        prev_active = driver.find_element(By.CSS_SELECTOR, "li.page-item.active a.page-link").text
                    except:
                        prev_active = None

                    driver.execute_script("arguments[0].click();", next_button)

                    def active_page_changed(drv):
                        try:
                            cur = drv.find_element(By.CSS_SELECTOR, "li.page-item.active a.page-link").text
                            return cur != prev_active
                        except:
                            return False

                    try:
                        wait.until(active_page_changed)
                    except:
                        # Fallback: wait for table rows to be present
                        try:
                            wait.until(EC.presence_of_element_located((By.CSS_SELECTOR, "table.dataTable tbody tr")))
                        except:
                            pass

                    time.sleep(1)  # small pause for table to fully update
                    page += 1
                    break  # Success, exit retry loop
                    
                except UnexpectedAlertPresentException:
                    print(f"Alert encountered, accepting it and retrying... (Attempt {retry_count + 1}/{max_retries})")
                    try:
                        alert = Alert(driver)
                        alert.accept()
                    except:
                        pass
                    time.sleep(2)
                    retry_count += 1
                    
                except Exception as e:
                    print(f"Error navigating to next page: {e}")
                    if retry_count < max_retries - 1:
                        print(f"Retrying... (Attempt {retry_count + 1}/{max_retries})")
                        time.sleep(2)
                        retry_count += 1
                    else:
                        print("Max retries reached, stopping scrape")
                        return
        
        # Save final data to Excel (convert checkpoint CSV to Excel using streaming if available)
        csv_checkpoint = os.path.splitext(output_file)[0] + ".csv"
        if csv_only:
            print(f"--csv-only set: leaving CSV checkpoint in place at {csv_checkpoint} and skipping Excel conversion.")
        else:
            if os.path.exists(csv_checkpoint):
                print(f"Converting checkpoint CSV to Excel: {csv_checkpoint} -> {output_file}")
                try:
                    csv_to_excel_stream(csv_checkpoint, output_file)
                except Exception as e:
                    print(f"Failed to convert CSV to Excel: {e}. Falling back to in-memory save.")
                    save_to_excel(data, output_file)
            else:
                print(f"Saving {len(data)} records to {output_file}")
                save_to_excel(data, output_file)
        
        print("Scraping completed successfully")
        
    except Exception as e:
        print(f"Error occurred: {str(e)}")
    finally:
        driver.quit()

if __name__ == "__main__":
    parser = argparse.ArgumentParser(description="NAFDAC Greenbook scraper")
    parser.add_argument("--start", type=int, help="Start page (overrides resume detection)")
    parser.add_argument("--end", type=int, default=876, help="End page")
    parser.add_argument("--driver", type=str, help="Full path to chromedriver executable to avoid auto-download")
    parser.add_argument("--file", type=str, default="nafdac_greenbook.xlsx", help="Output Excel file path")
    parser.add_argument("--no-headless", action="store_true", help="Run Chrome with visible UI for debugging")
    parser.add_argument("--debug", action="store_true", help="Enable debug JS dumps to stdout")
    parser.add_argument("--force-api", action="store_true", help="Run using HTTP API only (no Selenium) if possible")
    parser.add_argument("--csv-only", action="store_true", help="Only write/appends to CSV checkpoint and skip Excel conversion")
    args = parser.parse_args()

    # Determine start page precedence: CLI arg > START_PAGE env var > start_page.txt file > existing checkpoint
    default_start = None
    # CLI will override; but if not provided, check env/file
    env_val = os.environ.get('START_PAGE')
    if env_val:
        try:
            default_start = int(env_val)
            print(f"Using START_PAGE from environment: {default_start}")
        except Exception:
            default_start = None
    # start_page.txt takes next precedence
    if default_start is None and os.path.exists('start_page.txt'):
        try:
            with open('start_page.txt', 'r', encoding='utf-8') as f:
                txt = f.read().strip()
                default_start = int(txt)
                print(f"Using start_page from start_page.txt: {default_start}")
        except Exception:
            default_start = None

    # If still not provided, try to compute from existing CSV/XLSX checkpoints
    if default_start is None:
        computed = compute_start_page_from_files()
        if computed:
            default_start = computed
            print(f"Detected existing checkpoint; computed start page: {default_start}")

    start_arg = args.start if args.start is not None else default_start

    # Call scraper with csv_only flag
    scrape_greenbook(output_file=args.file, end_page=args.end, resume=True, driver_path=args.driver, start_page=start_arg, no_headless=args.no_headless, debug=args.debug, force_api=args.force_api, csv_only=args.csv_only)
